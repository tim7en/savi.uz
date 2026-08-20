"""What the economy looked like before each Federal Reserve decision.

The chapter has shown that the surprise at an announcement cannot be forecast.
That invites an obvious objection: the Fed is not capricious, it responds to
inflation and employment, and both are published in advance. If the reaction
function is known, why is the outcome not?

The answer separates two things the word "decision" runs together. What the
committee *does* is largely predictable from the data it watches. What the market
*learns* is the gap between that and what it had already priced. This measures
both on the same set of meetings, so the reader can see the first succeed and the
second fail rather than take it on assurance.

Publication lag is respected rather than assumed away. Monthly statistics are
taken only where the observation month ended at least forty-five days before the
meeting, which guarantees the figure had been released; market prices --
breakevens, credit spreads -- are taken from the prior session, when they were
quoted. Revisions are a known weakness: these are the current vintages, not the
numbers as first printed, and where the store holds first releases they begin too
late to cover the period.
"""

from __future__ import annotations

import argparse
import json
import math
import sqlite3
import statistics
from datetime import date, timedelta
from pathlib import Path

MONTHLY_LAG = 45


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--macro", type=Path, default=Path("data/macro/macro.db"))
    parser.add_argument("--events", type=Path, default=Path("data/macro/USMPD.xlsx"))
    parser.add_argument("--out", type=Path,
                        default=Path("out/report/chapter2_reaction.json"))
    return parser.parse_args(argv)


def series(connection, series_id):
    rows = connection.execute(
        "SELECT obs_date, value FROM observations WHERE series_id=? "
        "AND value IS NOT NULL ORDER BY obs_date", (series_id,)).fetchall()
    return [(d[:10], v) for d, v in rows]


def latest_before(data, cutoff):
    best = None
    for day, value in data:
        if day <= cutoff:
            best = (day, value)
        else:
            break
    return best


def year_change(data, cutoff):
    """Percentage change against the reading twelve months earlier."""
    now = latest_before(data, cutoff)
    if not now:
        return None
    then = latest_before(data, (date.fromisoformat(now[0])
                                - timedelta(days=365)).isoformat())
    if not then or not then[1]:
        return None
    return (now[1] / then[1] - 1) * 100


def meetings(path: Path):
    import openpyxl
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Monetary Events"]
    header = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    idx = {name: header.index(name) for name in ("Date", "Unscheduled", "SEP", "UST10Y")}
    out = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stamp = row[idx["Date"]]
        if stamp is None or row[idx["Unscheduled"]]:
            continue
        day = stamp.strftime("%Y-%m-%d")
        if day >= "2000-01-01":
            out.append({"date": day, "sep": bool(row[idx["SEP"]]),
                        "surprise": row[idx["UST10Y"]]})
    workbook.close()
    return sorted(out, key=lambda m: m["date"])


def main(argv=None):
    args = parse_args(argv)
    connection = sqlite3.connect(f"file:{args.macro}?mode=ro", uri=True)
    core = series(connection, "PCEPILFE")
    unemployment = series(connection, "UNRATE")
    payrolls = series(connection, "PAYEMS")
    balance = series(connection, "WALCL")
    breakeven = series(connection, "T5YIFR")
    spread = series(connection, "BAA10Y")
    funds = dict(series(connection, "DFF"))
    connection.close()
    funds_days = sorted(funds)

    rows = []
    for meeting in meetings(args.events):
        day = meeting["date"]
        monthly_cut = (date.fromisoformat(day) - timedelta(days=MONTHLY_LAG)).isoformat()
        previous = (date.fromisoformat(day) - timedelta(days=1)).isoformat()
        rate_now = latest_before([(d, funds[d]) for d in funds_days], previous)
        ahead = (date.fromisoformat(day) + timedelta(days=120)).isoformat()
        rate_later = latest_before([(d, funds[d]) for d in funds_days], ahead)
        jobs = latest_before(payrolls, monthly_cut)
        jobs_prior = (latest_before(payrolls,
                                    (date.fromisoformat(jobs[0])
                                     - timedelta(days=95)).isoformat())
                      if jobs else None)
        row = {
            "date": day, "sep": meeting["sep"], "surprise": meeting["surprise"],
            "core_pce": year_change(core, monthly_cut),
            "unemployment": (latest_before(unemployment, monthly_cut) or (None, None))[1],
            # PAYEMS is already in thousands of persons.
            "payroll_3m": ((jobs[1] - jobs_prior[1]) / 3
                           if jobs and jobs_prior else None),
            "balance_sheet": year_change(balance, previous),
            "breakeven": (latest_before(breakeven, previous) or (None, None))[1],
            "credit_spread": (latest_before(spread, previous) or (None, None))[1],
            "rate": rate_now[1] if rate_now else None,
            "rate_ahead": rate_later[1] if rate_later else None,
        }
        if row["rate"] is not None and row["rate_ahead"] is not None:
            row["rate_move"] = row["rate_ahead"] - row["rate"]
        rows.append(row)

    usable = [r for r in rows if r.get("rate_move") is not None]
    for row in usable:
        move = row["rate_move"]
        row["action"] = ("Raised" if move >= 0.20 else
                         "Cut" if move <= -0.20 else "Held")

    print(f"{len(usable)} meetings with a measurable subsequent rate path\n")
    fields = [("core_pce", "Core inflation, %yr"), ("unemployment", "Unemployment, %"),
              ("payroll_3m", "Jobs, 3m avg (000s)"),
              ("balance_sheet", "Balance sheet, %yr"),
              ("breakeven", "5y5y breakeven, %"),
              ("credit_spread", "Baa spread, pp")]
    print(f"  {'condition before the meeting':26s} " +
          "".join(f"{a:>12s}" for a in ("Raised", "Held", "Cut")))
    summary = {}
    for key, label in fields:
        cells, line = {}, ""
        for action in ("Raised", "Held", "Cut"):
            values = [r[key] for r in usable if r["action"] == action
                      and r[key] is not None]
            cells[action] = statistics.fmean(values) if values else None
            line += f"{cells[action]:>12.2f}" if values else f"{'—':>12s}"
        summary[key] = {"label": label, **cells}
        print(f"  {label:26s}" + line)
    counts = {a: sum(1 for r in usable if r["action"] == a)
              for a in ("Raised", "Held", "Cut")}
    print(f"  {'meetings':26s}" + "".join(f"{counts[a]:>12d}"
                                          for a in ("Raised", "Held", "Cut")))

    # In-sample R-squared rises with the number of predictors whether or not they
    # help, so both questions are asked out of sample: fitted on meetings before
    # 2017 and scored on those after, which is the same split the rest of this
    # work uses.
    predictors = [k for k, _ in fields]

    def solve(rows_in, target_key):
        data = [r for r in rows_in if r.get(target_key) is not None
                and all(r[k] is not None for k in predictors)]
        if len(data) < 30:
            return None, None
        X = [[1.0] + [r[k] for k in predictors] for r in data]
        y = [r[target_key] for r in data]
        columns = len(X[0])
        xtx = [[sum(row[a] * row[b] for row in X) for b in range(columns)]
               for a in range(columns)]
        xty = [sum(row[a] * y[i] for i, row in enumerate(X)) for a in range(columns)]
        for a in range(columns):
            xtx[a][a] += 1e-6
        for a in range(columns):
            pivot = max(range(a, columns), key=lambda r: abs(xtx[r][a]))
            xtx[a], xtx[pivot] = xtx[pivot], xtx[a]
            xty[a], xty[pivot] = xty[pivot], xty[a]
            for b in range(a + 1, columns):
                factor = xtx[b][a] / xtx[a][a]
                for k in range(a, columns):
                    xtx[b][k] -= factor * xtx[a][k]
                xty[b] -= factor * xty[a]
        beta = [0.0] * columns
        for a in reversed(range(columns)):
            beta[a] = (xty[a] - sum(xtx[a][b] * beta[b]
                                    for b in range(a + 1, columns))) / xtx[a][a]
        return beta, data

    def out_of_sample(target_key):
        train = [r for r in usable if r["date"] < "2017-01-01"]
        test = [r for r in usable if r["date"] >= "2017-01-01"]
        beta, fitted = solve(train, target_key)
        if beta is None:
            return None
        scored = [r for r in test if r.get(target_key) is not None
                  and all(r[k] is not None for k in predictors)]
        if len(scored) < 20:
            return None
        y = [r[target_key] for r in scored]
        pred = [beta[0] + sum(beta[i + 1] * r[k] for i, k in enumerate(predictors))
                for r in scored]
        mean = statistics.fmean([r[target_key] for r in fitted])
        ss_res = sum((y[i] - pred[i]) ** 2 for i in range(len(y)))
        ss_tot = sum((v - mean) ** 2 for v in y)
        correct = sum(1 for i in range(len(y))
                      if (y[i] > 0) == (pred[i] > 0)) / len(y)
        return {"n_train": len(fitted), "n_test": len(y),
                "r2": 1 - ss_res / ss_tot if ss_tot else None,
                "sign": correct}

    decision = out_of_sample("rate_move")
    surprise = out_of_sample("surprise")
    print(f"\nthe same six conditions, fitted before 2017 and scored after:")
    print(f"    what the committee DOES next   R2 {decision['r2']:>+7.3f}   "
          f"direction right {decision['sign']:.0%}   n = {decision['n_test']}")
    print(f"    what the market LEARNS         R2 {surprise['r2']:>+7.3f}   "
          f"direction right {surprise['sign']:.0%}   n = {surprise['n_test']}")

    report = {"summary": summary, "counts": counts, "decision": decision,
              "surprise": surprise, "rows": usable}
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(report, indent=1, default=str), encoding="utf-8")
    print(f"\n  wrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
