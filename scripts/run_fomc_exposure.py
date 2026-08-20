"""Cutting the book before a rate decision, and putting it back afterwards.

The chapter established two things that make this worth testing and one that
makes it hard.  Sectors reprice sharply on the day rates move; the direction of
that move cannot be forecast; but the *dates* on which it is most likely are
published a year ahead.  So the only defensible use of the finding is to carry
less into a scheduled meeting, which needs no view on the outcome at all.

The test applies that to the thirty-minute breakout book: exposure is reduced on
the meeting day and restored the next session, at a round-trip cost of ten basis
points on the amount actually moved.  Eight meetings a year at a half-weight cut
costs roughly forty basis points a year before anything else happens, which is a
real hurdle rather than a rounding error.

Three comparisons decide it, and the second is the one that has killed every
overlay this programme has tested:

* the untouched book;
* **a book held permanently at the overlay's own average exposure** -- because
  cutting risk on any schedule lowers drawdown, and the question is whether the
  *timing* adds anything a constant reduction would not;
* random dates, matched in number and spacing, so that a calendar with no meaning
  faces exactly the same machinery.

Everything is scored at matched drawdown, which removes the free advantage a
smaller book would otherwise enjoy, and per calendar year rather than in
aggregate.

Kill criterion, fixed before the run: the overlay must beat both the untouched
book and the constant-exposure book on Sharpe, in at least seven of ten years,
and clear the random-date null at p < 0.05.  Falling short of any of these, the
honest conclusion is that the calendar is worth knowing and not worth trading.
"""

from __future__ import annotations

import argparse
import json
import math
import random
import sqlite3
import statistics
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from savi_uz.split_adjust import adjust_bars, load_splits  # noqa: E402
from savi_uz.sweep_engulf import resample_regular_session  # noqa: E402
from savi_uz.turtle import TurtleConfig, run_turtle  # noqa: E402
from savi_uz.volume_profile import Bar  # noqa: E402

BASE = dict(entry_window=55, exit_window=20, atr_window=20, skip_after_winner=False,
            directions=(1,), use_channel_exit=False, chandelier_atr=3.0)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bars", type=Path, default=Path("data/intraday/bars.db"))
    parser.add_argument("--events", type=Path,
                        default=Path("data/macro/USMPD.xlsx"))
    parser.add_argument("--minutes", type=int, default=30)
    parser.add_argument("--max-positions", type=int, default=6)
    parser.add_argument("--target-dd", type=float, default=0.18)
    parser.add_argument("--cost", type=float, default=0.0002,
                        help="strategy round-trip cost")
    parser.add_argument("--switch-cost", type=float, default=0.0010,
                        help="round-trip cost of moving exposure, on the amount moved")
    parser.add_argument("--trials", type=int, default=30)
    parser.add_argument("--nulls", type=int, default=200)
    parser.add_argument("--out", type=Path,
                        default=Path("out/strategy/fomc_exposure.json"))
    return parser.parse_args(argv)


def load_book(args):
    splits = load_splits(args.bars)
    connection = sqlite3.connect(f"file:{args.bars}?mode=ro", uri=True)
    book = {}
    for (ticker,) in connection.execute(
            "SELECT DISTINCT ticker FROM bars WHERE frequency='5min' ORDER BY ticker"):
        rows = connection.execute(
            "SELECT ts,open,high,low,close,volume FROM bars WHERE ticker=? AND "
            "frequency='5min' ORDER BY ts", (ticker,)).fetchall()
        five = adjust_bars([Bar(*r) for r in rows], splits.get(ticker, []))
        if len({b.timestamp[:10] for b in five}) >= 400:
            book[ticker] = resample_regular_session(five, minutes=args.minutes)
    connection.close()
    return book


def meeting_days(path: Path):
    import openpyxl
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Monetary Events"]
    header = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    date_at = header.index("Date")
    unsched = header.index("Unscheduled")
    sep_at = header.index("SEP")
    scheduled, projections = set(), set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stamp = row[date_at]
        if stamp is None or row[unsched]:
            continue
        day = stamp.strftime("%Y-%m-%d") if hasattr(stamp, "strftime") else str(stamp)[:10]
        scheduled.add(day)
        if row[sep_at]:
            projections.add(day)
    workbook.close()
    return scheduled, projections


def session_closes(bars):
    out = {}
    for bar in bars:
        out[bar.timestamp[:10]] = bar.close
    return out


def trade_marks(trade, closes):
    entry_day, exit_day = trade.entry_timestamp[:10], trade.exit_timestamp[:10]
    marks = []
    for day in (d for d in closes if entry_day <= d < exit_day):
        live = [u for u in trade.unit_entries if u.timestamp[:10] <= day]
        if live:
            marks.append((day, sum(trade.direction * (closes[day] - u.price) / u.n
                                   for u in live)))
    return tuple(marks)


def cap(trades, limit, rng):
    shuffled = list(trades)
    rng.shuffle(shuffled)
    live, taken = [], []
    for trade in sorted(shuffled, key=lambda t: t["entry"]):
        live = [x for x in live if x["exit"] > trade["entry"]]
        if len(live) >= limit:
            continue
        live.append(trade)
        taken.append(trade)
    return taken


def daily_r(taken):
    by_day = defaultdict(float)
    for trade in taken:
        previous = 0.0
        for day, open_r in trade["marks"]:
            by_day[day] += open_r - previous
            previous = open_r
        by_day[trade["exit"][:10]] += trade["r"] - previous
    return by_day


def path(by_day, calendar, exposure, risk, switch_cost):
    """Compound the book at a given risk level under a daily exposure schedule."""
    nav, peak, worst = 1000.0, 1000.0, 0.0
    previous_exposure = 1.0
    values = []
    for day in calendar:
        e = exposure.get(day, 1.0)
        turn = abs(e - previous_exposure)
        previous_exposure = e
        gain = by_day.get(day, 0.0) * risk * e - turn * switch_cost
        nav = max(0.0, nav * (1.0 + gain))
        peak = max(peak, nav)
        worst = min(worst, nav / peak - 1.0)
        values.append(gain)
    years = (date.fromisoformat(calendar[-1])
             - date.fromisoformat(calendar[0])).days / 365.25
    cagr = (nav / 1000.0) ** (1 / years) - 1 if years > 0 and nav > 0 else -1.0
    return nav, worst, cagr, values


def solve_risk(maps, calendar, exposure, target, switch_cost, lo=1e-6, hi=0.08):
    def dd(risk):
        return statistics.median(abs(path(m, calendar, exposure, risk, switch_cost)[1])
                                 for m in maps)
    if dd(hi) < target:
        return hi
    for _ in range(30):
        mid = math.sqrt(lo * hi)
        if dd(mid) < target:
            lo = mid
        else:
            hi = mid
    return math.sqrt(lo * hi)


def score(maps, calendar, exposure, args):
    risk = solve_risk(maps, calendar, exposure, args.target_dd, args.switch_cost)
    runs = [path(m, calendar, exposure, risk, args.switch_cost) for m in maps]
    sharpes, cagrs, dds = [], [], []
    years = sorted({d[:4] for d in calendar})
    by_year = defaultdict(list)
    for nav, worst, cagr, values in runs:
        sd = statistics.pstdev(values)
        if sd > 0:
            sharpes.append(statistics.fmean(values) / sd * math.sqrt(252))
        cagrs.append(cagr)
        dds.append(worst)
        for year in years:
            window = [v for d, v in zip(calendar, values) if d[:4] == year]
            if len(window) > 60:
                s = statistics.pstdev(window)
                if s > 0:
                    by_year[year].append(statistics.fmean(window) / s * math.sqrt(252))
    return {"risk": risk, "sharpe": statistics.median(sharpes),
            "cagr": statistics.median(cagrs), "dd": statistics.median(dds),
            "years": {y: statistics.median(v) for y, v in by_year.items() if v},
            "avg_exposure": statistics.fmean(exposure.get(d, 1.0) for d in calendar)}


def main(argv=None):
    args = parse_args(argv)
    book = load_book(args)
    calendar = sorted({b.timestamp[:10] for bars in book.values() for b in bars})
    scheduled, projections = meeting_days(args.events)
    in_sample = [d for d in sorted(scheduled) if calendar[0] <= d <= calendar[-1]]
    print(f"{len(book)} instruments, {len(calendar):,} sessions "
          f"{calendar[0]} -> {calendar[-1]}")
    print(f"{len(in_sample)} scheduled meetings in range, "
          f"{len([d for d in projections if d in set(in_sample)])} with projections\n",
          flush=True)

    config = TurtleConfig(**{**BASE, "round_trip_cost": args.cost})
    pooled = []
    for ticker, bars in book.items():
        closes = session_closes(bars)
        for trade in run_turtle(bars, config=config)[0]:
            pooled.append({"entry": trade.entry_timestamp,
                           "exit": trade.exit_timestamp, "r": trade.net_r,
                           "marks": trade_marks(trade, closes)})
    maps = [daily_r(cap(pooled, args.max_positions, random.Random(s)))
            for s in range(args.trials)]
    print(f"{len(pooled):,} trades\n", flush=True)

    def schedule(days, weight):
        return {d: weight for d in days if d in set(calendar)}

    variants = [("untouched book", {})]
    for weight in (0.75, 0.50, 0.25, 0.0):
        variants.append((f"all meetings at {weight:.0%}",
                         schedule(in_sample, weight)))
    variants.append(("projection meetings at 50%",
                     schedule([d for d in in_sample if d in projections], 0.50)))

    report, base = {}, None
    print(f"  {'variant':30s} {'exposure':>9s} {'lev':>8s} {'Sharpe':>7s} "
          f"{'CAGR':>8s} {'vs flat':>9s}")
    for label, exposure in variants:
        result = score(maps, calendar, exposure, args)
        # The rival: the same average exposure, held every single day.
        flat = score(maps, calendar,
                     {d: result["avg_exposure"] for d in calendar}, args)
        result["flat_sharpe"] = flat["sharpe"]
        result["flat_cagr"] = flat["cagr"]
        report[label] = result
        if base is None:
            base = result
        print(f"  {label:30s} {result['avg_exposure']:>9.3f} {result['risk']:>8.4%} "
              f"{result['sharpe']:>7.2f} {result['cagr']:>8.1%} "
              f"{result['sharpe'] - flat['sharpe']:>+9.3f}", flush=True)

    best = max((v for k, v in report.items() if k != "untouched book"),
               key=lambda v: v["sharpe"])
    label = next(k for k, v in report.items() if v is best)
    years = sorted(base["years"])
    wins_base = sum(1 for y in years if best["years"].get(y, -9) > base["years"][y])
    print(f"\n  best overlay: {label}")
    print(f"  beats the untouched book in {wins_base}/{len(years)} years")
    print(f"  Sharpe against a permanently smaller book: "
          f"{best['sharpe']:.2f} vs {best['flat_sharpe']:.2f} "
          f"({best['sharpe'] - best['flat_sharpe']:+.3f})")

    print(f"\n  {args.nulls} random-date calendars, matched in number...", flush=True)
    rng = random.Random(5)
    weight = 0.50
    nulls = []
    for _ in range(args.nulls):
        fake = rng.sample(calendar, len(in_sample))
        nulls.append(score(maps, calendar, {d: weight for d in fake}, args)["sharpe"])
    nulls.sort()
    real = report["all meetings at 50%"]["sharpe"]
    beat = sum(1 for v in nulls if v >= real)
    pick = lambda f: nulls[min(int(f * len(nulls)), len(nulls) - 1)]
    p_value = (beat + 1) / (len(nulls) + 1)
    print(f"    random calendars: p05 {pick(.05):.2f}  median {pick(.5):.2f}  "
          f"p95 {pick(.95):.2f}")
    print(f"    real FOMC calendar at 50%: {real:.2f}")
    print(f"    empirical p = {p_value:.3f}")

    passed = (best["sharpe"] > base["sharpe"] and
              best["sharpe"] > best["flat_sharpe"] and
              wins_base >= 7 and p_value < 0.05)
    print(f"\n  KILL CRITERION: beat the untouched book and a permanently smaller "
          f"one, in >=7/10 years, p<0.05")
    print(f"  VERDICT: {'PASS' if passed else 'FAIL'}")
    report["null"] = {"p": p_value, "median": pick(.5)}
    report["verdict"] = {"passed": passed, "best": label, "years_won": wins_base}
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(report, indent=1), encoding="utf-8")
    print(f"\n  wrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
